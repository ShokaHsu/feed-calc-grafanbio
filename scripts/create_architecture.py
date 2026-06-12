from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
FONT = "Arial"

# ── Shared helpers ──────────────────────────────────────────────────────────

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def center(wrap=True):  return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def left(wrap=True):    return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)

def hdr(ws, row, cols_values, bg="1F3864", fg="FFFFFF", height=22):
    for col, val in enumerate(cols_values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name=FONT, bold=True, size=10, color=fg)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin()
    ws.row_dimensions[row].height = height

def cell(ws, row, col, val, bold=False, size=10, color="000000",
         bg="FFFFFF", h_align="left", wrap=True, height=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name=FONT, bold=bold, size=size, color=color)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
    c.border = thin()
    if height:
        ws.row_dimensions[row].height = height
    return c

def widths(ws, *pairs):
    for letter, w in pairs:
        ws.column_dimensions[letter].width = w

def new_sheet(name):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    return ws

def title_row(ws, text, cols=8, bg="1F3864"):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=FONT, bold=True, size=14, color="FFFFFF")
    c.fill = fill(bg)
    c.alignment = center()
    ws.row_dimensions[1].height = 32

# Layer palette
L = {
    "Shell":    ("BDD7EE", "1F3864"),   # light blue
    "UI":       ("DAEEF3", "17375E"),
    "State":    ("E2EFDA", "375623"),   # light green
    "APIclient":("FFF2CC", "7F6000"),   # yellow
    "Auth":     ("FCE4D6", "843C0C"),   # orange
    "Router":   ("F4CCFF", "4A086B"),   # purple
    "BizLogic": ("FFD7E5", "C00000"),   # pink/red
    "ORM":      ("D9D9D9", "404040"),   # grey
    "DB":       ("2F4858", "FFFFFF"),   # dark teal
    "Config":   ("F2F2F2", "595959"),   # light grey (cross-cutting)
}

# ════════════════════════════════════════════════════════════════════════════
# Sheet 1: Overview (ASCII architecture + summary table)
# ════════════════════════════════════════════════════════════════════════════
ws_ov = wb.active
ws_ov.title = "Overview"
ws_ov.sheet_view.showGridLines = False

ws_ov.merge_cells("A1:F1")
c = ws_ov.cell(row=1, column=1,
               value="Feed Formula Calculator — Layer Architecture Overview")
c.font = Font(name=FONT, bold=True, size=16, color="FFFFFF")
c.fill = fill("1F3864")
c.alignment = center()
ws_ov.row_dimensions[1].height = 35

# Sub-title
ws_ov.merge_cells("A2:F2")
c = ws_ov.cell(row=2, column=1,
               value="3-platform (Desktop · Cloud · Mobile-planned) | Django REST + Vue 3 + Tauri v2")
c.font = Font(name=FONT, italic=True, size=10, color="595959")
c.alignment = center()
ws_ov.row_dimensions[2].height = 18

# ASCII diagram block
diagram = """\
  ┌──────────────────────────────────────────────────────────────────────────────┐
  │                          LAYER 1 · SHELL / HOST                              │
  │   Tauri v2 WebView (Desktop/Win+Mac)   Browser (Cloud)   Capacitor (Mobile) │
  └──────────────────────────────┬───────────────────────────────────────────────┘
                                  │  renders
  ┌──────────────────────────────▼───────────────────────────────────────────────┐
  │                       LAYER 2 · UI COMPONENTS (Vue 3)                        │
  │  FormulaCalculator · IngredientList · NutrientAnalysisPanel · FormulaExporter│
  │  SwinePanel · PoultryPanel · GeneralPanel · NutrientRow · IngredientPicker   │
  └──────────────────────────────┬───────────────────────────────────────────────┘
                                  │  reads / commits
  ┌──────────────────────────────▼───────────────────────────────────────────────┐
  │                    LAYER 3 · STATE MANAGEMENT (Pinia)                        │
  │   useFormulaStore · useIngredientStore · usePreferencesStore                 │
  │   Composables: useGroupStatus                                                 │
  └──────────────────────────────┬───────────────────────────────────────────────┘
                                  │  HTTP via
  ┌──────────────────────────────▼───────────────────────────────────────────────┐
  │                       LAYER 4 · API CLIENT (Axios)                           │
  │   baseURL: .env / .env.desktop / .env.production                             │
  │   Token injection · 15 s timeout · 3-retry on startup                        │
  └──────────────────────────────┬───────────────────────────────────────────────┘
                                  │ REST / JSON
               ┌──────────────────┴──────────────────┐
               │  localhost:8042 (Desktop)             │  Railway HTTPS (Cloud)
               │  Waitress sidecar                     │  Gunicorn / WhiteNoise
  ┌────────────▼──────────────────────────────────────▼────────────────────────┐
  │                    LAYER 5 · AUTHENTICATION                                  │
  │   Desktop: StandaloneBypassAuth (auto local_admin, no token)                 │
  │   Cloud:   DRF Token Auth via Djoser  (/api/auth/token/login/)               │
  └──────────────────────────────┬──────────────────────────────────────────────┘
                                  │ passes request to
  ┌──────────────────────────────▼──────────────────────────────────────────────┐
  │                    LAYER 6 · API / ROUTING (DRF Views + Serializers)         │
  │   /api/ingredients/   /api/formulas/   /api/formulas/import/                 │
  │   /api/standards/     /api/auth/       /api/preferences/                     │
  └──────────────────────────────┬──────────────────────────────────────────────┘
                                  │ calls
  ┌──────────────────────────────▼──────────────────────────────────────────────┐
  │                    LAYER 7 · BUSINESS LOGIC                                  │
  │   FormulaService.calculate_and_save()   _parse_formula_csv()                 │
  │   IsOwnerOrReadOnly permission                                                │
  └──────────────────────────────┬──────────────────────────────────────────────┘
                                  │ queries via
  ┌──────────────────────────────▼──────────────────────────────────────────────┐
  │                    LAYER 8 · DATA ACCESS (Django ORM)                        │
  │   Formula · FormulaItem · Ingredient · NutrientStandard                      │
  │   User · Organization · Customer · UserPreference                             │
  └──────────────────────────────┬──────────────────────────────────────────────┘
                                  │ SQL
  ┌──────────────────────────────▼──────────────────────────────────────────────┐
  │                    LAYER 9 · DATABASE                                        │
  │   SQLite  — Desktop & local dev  (APPDATA\\FeedCalc\\db.sqlite3)              │
  │   PostgreSQL — Cloud production  (Railway DATABASE_URL)                      │
  └──────────────────────────────────────────────────────────────────────────────┘

  ════════════════  CROSS-CUTTING CONCERNS  ════════════════
  Configuration:  settings/base.py · desktop.py · prod.py · .env files
  Packaging:      PyInstaller (sidecar .exe) · Vite (frontend bundle) · Tauri build
  Seeding:        crawled_seed.json · import_ingredients · load_nrc_json · run_api._load_seed_fixture()
  CORS:           base.py CORS_ALLOWED_ORIGINS + x-client-mode header (desktop.py)
"""

ws_ov.merge_cells("A4:F38")
c = ws_ov.cell(row=4, column=1, value=diagram)
c.font = Font(name="Courier New", size=9, color="1F1F1F")
c.fill = fill("F8F8F8")
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
c.border = thin()
ws_ov.row_dimensions[4].height = 15
for r in range(5, 39):
    ws_ov.row_dimensions[r].height = 15

widths(ws_ov, ("A", 20), ("B", 20), ("C", 20), ("D", 20), ("E", 20), ("F", 20))

# ════════════════════════════════════════════════════════════════════════════
# Sheet 2: Layer Definitions
# ════════════════════════════════════════════════════════════════════════════
ws_ld = new_sheet("Layer Definitions")
title_row(ws_ld, "Layer Definitions — Responsibilities, Technologies & Interfaces", cols=9)

hdr(ws_ld, 2, ["#", "Layer Name", "Tier", "Platform",
               "Responsibility", "Key Files / Modules", "Technology", "Interfaces With", "Test Strategy"])
widths(ws_ld,
    ("A", 5), ("B", 22), ("C", 12), ("D", 14),
    ("E", 45), ("F", 40), ("G", 25), ("H", 28), ("I", 35))

layers = [
    (1, "Shell / Host",       "Client",   "All",
     "Wraps the frontend bundle. Manages window chrome, OS integration, "
     "and IPC (Tauri). Determines which API base URL is used.",
     "src-tauri/tauri.conf.json\nsrc-tauri/src/main.rs",
     "Tauri v2 (Desktop)\nBrowser (Cloud)\nCapacitor (Mobile, planned)",
     "→ UI Components (renders)\n→ Layer 4 (environment vars)",
     "Manual / E2E (Playwright on Cloud, tauri-driver on Desktop)"),

    (2, "UI Components",      "Frontend", "All",
     "Renders the user interface. Receives state from stores (read-only) "
     "and dispatches actions. Splits by concern: formula editing, "
     "ingredient browsing, nutrient analysis panels.",
     "src/components/FormulaCalculator.vue\n"
     "src/components/formula/FormulaTable.vue\n"
     "src/components/formula/NutrientAnalysisPanel.vue\n"
     "src/components/panels/SwinePanel.vue\n"
     "src/components/panels/PoultryPanel.vue\n"
     "src/components/panels/GeneralPanel.vue\n"
     "src/components/FormulaExporter.vue\n"
     "src/components/IngredientList.vue",
     "Vue 3 SFCs\nElement Plus\nVite",
     "← Layer 3 (reads store)\n→ Layer 3 (dispatches actions)",
     "Vitest component tests\nPlaywright E2E (golden path)"),

    (3, "State Management",   "Frontend", "All",
     "Centralises async data fetching and reactive state. "
     "Stores own their slice of server state and expose computed getters. "
     "Composables encapsulate reusable reactive logic shared across panels.",
     "src/stores/useFormulaStore.js\n"
     "src/stores/useIngredientStore.js\n"
     "src/stores/usePreferencesStore.js\n"
     "src/composables/useGroupStatus.js",
     "Pinia\nVue 3 Composition API",
     "← Layer 2 (consumed by components)\n→ Layer 4 (issues API calls)",
     "Vitest unit tests\n(mock Axios, test store actions)"),

    (4, "API Client",         "Frontend", "All",
     "Wraps all HTTP communication. Injects auth token header, "
     "configures timeout (15 s), handles startup retry (3× at 2 s gap). "
     "Base URL switches per environment via Vite .env files.",
     "src/api/ (axios instance)\n"
     "feed_calc_frontend/.env\n"
     "feed_calc_frontend/.env.desktop\n"
     "feed_calc_frontend/.env.production",
     "Axios\nVite env variables",
     "← Layer 3 (called by stores)\n→ Layer 5 (HTTP to backend)",
     "Vitest with axios-mock-adapter\nor MSW"),

    (5, "Authentication",     "Backend",  "All",
     "Desktop: StandaloneBypassAuth auto-authenticates every request "
     "as 'local_admin' (superuser) — no token required. "
     "Cloud: Djoser-managed token auth; POST /api/auth/token/login/ "
     "returns auth_token; subsequent requests carry 'Token <token>' header.",
     "feed_calc_backend/accounts/authentication.py\n"
     "feed_calc_backend/config/settings/desktop.py (StandaloneBypassAuth inline)\n"
     "feed_calc_backend/config/settings/base.py (DEFAULT_AUTHENTICATION_CLASSES)",
     "Django REST Framework\nDjoser\nDRF Token Auth",
     "← Layer 4 (receives HTTP)\n→ Layer 6 (attaches user to request)",
     "pytest-django\n(test both auth paths with separate settings)"),

    (6, "API / Routing",      "Backend",  "All",
     "URL routing + DRF generic views + serializers. Validates request "
     "data, enforces object-level permissions (IsOwnerOrReadOnly), "
     "and serialises ORM objects to JSON.",
     "feed_calc_backend/formulas/views.py\n"
     "feed_calc_backend/formulas/serializers.py\n"
     "feed_calc_backend/ingredients/views.py\n"
     "feed_calc_backend/standards/views.py\n"
     "feed_calc_backend/accounts/views.py\n"
     "feed_calc_backend/*/urls.py",
     "Django REST Framework\ngeneric views\nModelSerializer",
     "← Layer 5 (authenticated request)\n→ Layer 7 (delegates business logic)\n→ Layer 8 (ORM queries)",
     "pytest-django API tests\n(explicit status + body assertions)"),

    (7, "Business Logic",     "Backend",  "All",
     "Domain-specific computation isolated from HTTP concerns. "
     "FormulaService recalculates cost snapshot after any item change. "
     "_parse_formula_csv parses the CSV export format for import. "
     "IsOwnerOrReadOnly enforces row-level write permissions.",
     "feed_calc_backend/formulas/services.py (FormulaService)\n"
     "feed_calc_backend/formulas/views.py (_parse_formula_csv)\n"
     "feed_calc_backend/formulas/views.py (IsOwnerOrReadOnly)",
     "Pure Python",
     "← Layer 6 (called by views)\n→ Layer 8 (reads/writes ORM)",
     "pytest unit tests\n(no HTTP, no DB fixtures needed\nfor CSV parser)"),

    (8, "Data Access (ORM)",  "Backend",  "All",
     "Django models define schema and business rules (constraints, "
     "cascades, ordering). QuerySets provide the only interface to the DB. "
     "Migrations track schema evolution across both SQLite and PostgreSQL.",
     "feed_calc_backend/formulas/models.py\n"
     "feed_calc_backend/ingredients/models.py\n"
     "feed_calc_backend/standards/models.py\n"
     "feed_calc_backend/accounts/models.py\n"
     "feed_calc_backend/*/migrations/",
     "Django ORM\ndjango.db.models",
     "← Layer 7 (business logic)\n→ Layer 9 (SQL to DB)",
     "pytest-django\n@pytest.mark.django_db\nfactory_boy fixtures"),

    (9, "Database",           "Infra",    "All",
     "Persistent storage. SQLite used for Desktop (file in APPDATA) "
     "and local dev. PostgreSQL used in cloud production (Railway). "
     "Schema managed entirely through Django migrations — no raw SQL.",
     "APPDATA\\FeedCalc\\db.sqlite3 (desktop)\n"
     "Railway DATABASE_URL (cloud/prod)\n"
     "feed_calc_backend/db.sqlite3 (dev fallback)",
     "SQLite 3 (Desktop / Dev)\nPostgreSQL 15 (Cloud)",
     "← Layer 8 (Django ORM)\n(no direct access from any other layer)",
     "pytest-django uses\nin-memory SQLite by default"),
]

LAYER_BG = [
    L["Shell"][0], L["UI"][0], L["State"][0], L["APIclient"][0],
    L["Auth"][0],  L["Router"][0], L["BizLogic"][0], L["ORM"][0], L["DB"][0],
]
LAYER_FG = [
    L["Shell"][1], L["UI"][1], L["State"][1], L["APIclient"][1],
    L["Auth"][1],  L["Router"][1], L["BizLogic"][1], L["ORM"][1], L["DB"][1],
]

for i, row_data in enumerate(layers, start=3):
    r = i
    bg = LAYER_BG[i-3]
    fg = LAYER_FG[i-3]
    ws_ld.row_dimensions[r].height = 80
    for col, val in enumerate(row_data, 1):
        c = ws_ld.cell(row=r, column=col, value=str(val))
        c.border = thin()
        c.alignment = Alignment(horizontal="left" if col > 1 else "center",
                                vertical="center", wrap_text=True)
        if col in (1, 2):
            c.font = Font(name=FONT, bold=True, size=10, color=fg)
            c.fill = fill(bg)
        else:
            c.font = Font(name=FONT, size=9, color="1F1F1F")
            c.fill = fill("FAFAFA" if i % 2 == 0 else "FFFFFF")

# Cross-cutting section
r = len(layers) + 3
ws_ld.row_dimensions[r].height = 22
hdr(ws_ld, r, ["Cross-Cutting Concerns", "", "", "", "", "", "", "", ""],
    bg=L["Config"][1], fg=L["Config"][0])
ws_ld.merge_cells(f"A{r}:I{r}")

cross = [
    ("Config / Settings", "Cross-cutting", "All",
     "Environment-specific Django settings. base.py shared by all. "
     "desktop.py overrides auth + DB path + CORS. prod.py adds PostgreSQL + HTTPS.",
     "config/settings/base.py\nconfig/settings/desktop.py\nconfig/settings/prod.py",
     "Django settings split", "All layers", "—"),
    ("Build & Packaging", "Cross-cutting", "All",
     "PyInstaller bundles Django + Waitress into feed_calc_api.exe. "
     "Vite builds the frontend bundle. Tauri packages both into a native installer. "
     "djoser.urls guarded with sys.frozen check to prevent PyInstaller crash.",
     "feed_calc_backend/feed_calc_api.spec\n"
     "feed_calc_frontend/src-tauri/tauri.conf.json\n"
     "scripts/build-desktop.ps1",
     "PyInstaller · Vite · Tauri v2", "Layers 1, 4, 5, 9", "Manual build test"),
    ("Data Seeding", "Cross-cutting", "Desktop",
     "On first desktop launch, run_api._load_seed_fixture() loads "
     "ingredients/fixtures/crawled_seed.json if DB is empty. "
     "Management commands (import_ingredients, load_nrc_json) used at build time.",
     "feed_calc_backend/run_api.py\n"
     "feed_calc_backend/ingredients/fixtures/crawled_seed.json\n"
     "feed_calc_backend/ingredients/management/commands/",
     "Django fixtures · management commands", "Layers 8, 9", "Integration test (P4-002, P4-003)"),
    ("CORS", "Cross-cutting", "All",
     "django-cors-headers configured in base.py. Desktop adds 'x-client-mode' "
     "to CORS_ALLOW_HEADERS. Preflight OPTIONS must succeed before any state-changing request.",
     "config/settings/base.py (CORS_ALLOWED_ORIGINS)\n"
     "config/settings/desktop.py (CORS_ALLOW_HEADERS)",
     "django-cors-headers", "Layers 4, 5", "Integration test (P4-006)"),
]

for j, cd in enumerate(cross, start=r+1):
    ws_ld.row_dimensions[j].height = 60
    for col, val in enumerate(("—",) + cd, 1):
        c2 = ws_ld.cell(row=j, column=col, value=str(val))
        c2.font = Font(name=FONT, size=9, italic=(col == 1))
        c2.fill = fill("F2F2F2" if j % 2 == 0 else "FAFAFA")
        c2.border = thin()
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ════════════════════════════════════════════════════════════════════════════
# Sheet 3: Component Map
# ════════════════════════════════════════════════════════════════════════════
ws_cm = new_sheet("Component Map")
title_row(ws_cm, "Component Map — Every Significant Module Assigned to Its Layer", cols=7)

hdr(ws_cm, 2, ["Layer", "File / Module", "Class / Function", "Platform", "Responsibility", "Key Dependencies", "Test File"])
widths(ws_cm, ("A", 22), ("B", 45), ("C", 30), ("D", 12), ("E", 40), ("F", 35), ("G", 25))

components = [
    # Shell
    ("1 · Shell", "src-tauri/tauri.conf.json", "beforeDevCommand / beforeBuildCommand", "Desktop",
     "Launches npm dev:desktop or build:desktop before Tauri window opens",
     "npm scripts, .env.desktop", "Manual"),
    ("1 · Shell", "src-tauri/src/main.rs", "main()", "Desktop",
     "Tauri app entry; sets up WebView and IPC plugins",
     "Tauri v2 runtime", "Manual"),

    # UI
    ("2 · UI", "src/components/FormulaCalculator.vue", "FormulaCalculator", "All",
     "Root formula editing view: ingredient picker + formula table + analysis panel",
     "useFormulaStore, useIngredientStore", "E2E"),
    ("2 · UI", "src/components/formula/FormulaTable.vue", "FormulaTable", "All",
     "Editable table of formula items with amount inputs and cost display",
     "useFormulaStore", "Vitest component"),
    ("2 · UI", "src/components/formula/NutrientAnalysisPanel.vue", "NutrientAnalysisPanel", "All",
     "Tabs container for Swine/Poultry/General panels + mode toggle",
     "SwinePanel, PoultryPanel, GeneralPanel", "Vitest component"),
    ("2 · UI", "src/components/panels/SwinePanel.vue", "SwinePanel", "All",
     "Renders pig-specific nutrient groups with status badges",
     "useGroupStatus, NutrientRow", "Vitest component"),
    ("2 · UI", "src/components/panels/PoultryPanel.vue", "PoultryPanel", "All",
     "Renders poultry-specific nutrient groups with status badges",
     "useGroupStatus, NutrientRow", "Vitest component"),
    ("2 · UI", "src/components/panels/GeneralPanel.vue", "GeneralPanel", "All",
     "Renders species-agnostic nutrients (minerals, vitamins, fatty acids)",
     "useGroupStatus, NutrientRow", "Vitest component"),
    ("2 · UI", "src/components/panels/NutrientRow.vue", "NutrientRow", "All",
     "Single nutrient row: label, progress bar, value vs target",
     "Props only (no store)", "Vitest component"),
    ("2 · UI", "src/components/FormulaExporter.vue", "FormulaExporter", "All",
     "Generates and downloads CSV export of a formula",
     "useFormulaStore", "Vitest (CSV content)"),
    ("2 · UI", "src/components/formula/IngredientPicker.vue", "IngredientPicker", "All",
     "Search + select UI for adding ingredients to formula",
     "useIngredientStore", "Vitest component"),
    ("2 · UI", "src/components/IngredientList.vue", "IngredientList", "All",
     "Browse/search all ingredients with filter and sort",
     "useIngredientStore", "E2E"),

    # State
    ("3 · State", "src/stores/useFormulaStore.js", "useFormulaStore", "All",
     "Owns formula list, current formula, items. Wraps formula CRUD + cost snapshot.",
     "Axios API client, useIngredientStore", "Vitest (store)"),
    ("3 · State", "src/stores/useIngredientStore.js", "useIngredientStore", "All",
     "Paginated ingredient list, search state, ingredient detail cache",
     "Axios API client", "Vitest (store)"),
    ("3 · State", "src/stores/usePreferencesStore.js", "usePreferencesStore", "All",
     "Persists nutrient_display_mode (percent vs absolute) via POST upsert",
     "Axios API client", "Vitest (store)"),
    ("3 · State", "src/composables/useGroupStatus.js", "useGroupStatus()", "All",
     "Takes [{current, target, isMax}] → {status, count}. Used by all three panels.",
     "None (pure function)", "Vitest unit — P3-001 to P3-004"),

    # API Client
    ("4 · API Client", "src/api/ (axios instance)", "axiosInstance", "All",
     "Configured Axios with baseURL, 15 s timeout, token interceptor",
     ".env files (VITE_API_BASE)", "Vitest + axios-mock-adapter"),
    ("4 · API Client", "feed_calc_frontend/.env.desktop", "VITE_API_BASE", "Desktop",
     "Points to http://127.0.0.1:8042/api — must never be changed to 8000",
     "run_api.py hardcodes port 8042", "Build-time check"),

    # Auth
    ("5 · Auth", "accounts/authentication.py", "DesktopStandaloneAuthentication", "Desktop",
     "Header-gated bypass: only triggers on 'Bearer standalone-admin'. Creates desktop_default_user (non-admin).",
     "Django AUTH_USER_MODEL", "pytest P2-004"),
    ("5 · Auth", "config/settings/desktop.py", "StandaloneBypassAuth (inline)", "Desktop",
     "No header check — always auto-authenticates as local_admin (superuser). "
     "INCONSISTENCY: different from DesktopStandaloneAuthentication in accounts/authentication.py",
     "Django REST Framework", "pytest P2-001 to P2-003"),
    ("5 · Auth", "config/settings/base.py", "DEFAULT_AUTHENTICATION_CLASSES", "Cloud",
     "DRF Token auth + JWT SimpleJWT. Djoser manages user CRUD endpoints.",
     "djoser, rest_framework.authtoken", "pytest P1-001 to P1-005"),

    # API/Routing
    ("6 · API", "formulas/views.py", "FormulaListView", "All",
     "GET list (all formulas) + POST create. Injects created_by=request.user.",
     "FormulaSerializer, FormulaService", "pytest P1-011"),
    ("6 · API", "formulas/views.py", "FormulaDetailView", "All",
     "GET/PUT/DELETE single formula. IsOwnerOrReadOnly enforces write restriction.",
     "FormulaSerializer, IsOwnerOrReadOnly", "pytest P1-012 to P1-016"),
    ("6 · API", "formulas/views.py", "FormulaImportView", "All",
     "POST multipart CSV → parse → match ingredients → create formula + items",
     "_parse_formula_csv, FormulaService, IngredientSerializer", "pytest P1-018 to P1-021"),
    ("6 · API", "ingredients/views.py", "IngredientViewSet", "All",
     "CRUD for ingredients. Filters by is_public=True OR created_by=user.",
     "IngredientSerializer", "pytest P1-006 to P1-010"),
    ("6 · API", "standards/views.py", "NutrientStandardViewSet", "All",
     "Read-only list of NRC standards (requirements + stages)",
     "StandardSerializer", "pytest P1-022, P1-023"),
    ("6 · API", "accounts/views.py", "UserPreferenceView", "All",
     "POST upsert of UserPreference (nutrient_display_mode, favorites)",
     "UserPreferenceSerializer", "pytest P2-005, P2-006"),

    # Business Logic
    ("7 · BizLogic", "formulas/services.py", "FormulaService.calculate_and_save()", "All",
     "Iterates FormulaItems, sums cost*weight, updates batch_size/total_cost/cost_per_kg snapshot on Formula.",
     "Formula, FormulaItem, Ingredient.cost_per_kg_twd", "pytest P0-001 to P0-004"),
    ("7 · BizLogic", "formulas/views.py", "_parse_formula_csv()", "All",
     "Pure function: bytes → (meta dict, [(name, amount_kg)]). Handles BOM, header rows, bad floats.",
     "csv, io (stdlib only)", "pytest P0-005 to P0-009"),
    ("7 · BizLogic", "formulas/views.py", "IsOwnerOrReadOnly", "All",
     "SAFE_METHODS allowed for any authenticated user; mutating methods only for created_by==user.",
     "rest_framework.permissions", "pytest P1-014, P1-016"),

    # ORM
    ("8 · ORM", "formulas/models.py", "Formula / FormulaItem", "All",
     "Formula: name, standard FK (SET_NULL), customer FK (SET_NULL), cost snapshot fields. "
     "FormulaItem: unique_together(formula, ingredient).",
     "NutrientStandard, Ingredient, Customer", "pytest @django_db"),
    ("8 · ORM", "ingredients/models.py", "Ingredient", "All",
     "80+ nutrient fields. BUG: phytate_P_g_per_kg line 87 uses ':' not '=' (missing from DB).",
     "AUTH_USER_MODEL", "pytest P5-001 (pinned fail)"),
    ("8 · ORM", "standards/models.py", "NutrientStandard", "All",
     "NRC min/max requirements per species+stage. Read-only via API.",
     "—", "pytest @django_db"),
    ("8 · ORM", "accounts/models.py", "User / UserPreference", "All",
     "AbstractUser + tier + org. UserPreference: nutrient_display_mode CharField.",
     "Organization, Customer", "pytest @django_db"),

    # DB
    ("9 · DB", "APPDATA/FeedCalc/db.sqlite3", "SQLite", "Desktop",
     "Desktop runtime DB. Created fresh on first launch; seeded via _load_seed_fixture().",
     "run_api.py, crawled_seed.json", "pytest with SQLite in-memory"),
    ("9 · DB", "Railway DATABASE_URL", "PostgreSQL 15", "Cloud",
     "Production DB on Railway. Managed via Django migrations only.",
     "config/settings/prod.py", "CI pipeline (separate)"),
]

LAYER_COLORS = {
    "1 · Shell":    (L["Shell"][0],    L["Shell"][1]),
    "2 · UI":       (L["UI"][0],       L["UI"][1]),
    "3 · State":    (L["State"][0],    L["State"][1]),
    "4 · API Client":(L["APIclient"][0], L["APIclient"][1]),
    "5 · Auth":     (L["Auth"][0],     L["Auth"][1]),
    "6 · API":      (L["Router"][0],   L["Router"][1]),
    "7 · BizLogic": (L["BizLogic"][0], L["BizLogic"][1]),
    "8 · ORM":      (L["ORM"][0],      L["ORM"][1]),
    "9 · DB":       (L["DB"][0],       L["DB"][1]),
}

for i, row_data in enumerate(components, start=3):
    r = i
    ws_cm.row_dimensions[r].height = 55
    layer_key = row_data[0]
    bg_l, fg_l = LAYER_COLORS.get(layer_key, ("FFFFFF", "000000"))
    row_bg = "F9F9F9" if i % 2 == 0 else "FFFFFF"
    for col, val in enumerate(row_data, 1):
        c = ws_cm.cell(row=r, column=col, value=val)
        c.border = thin()
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if col == 1:
            c.font = Font(name=FONT, bold=True, size=9, color=fg_l)
            c.fill = fill(bg_l)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            c.font = Font(name=FONT, size=9, color="1F1F1F")
            c.fill = fill(row_bg)

# ════════════════════════════════════════════════════════════════════════════
# Sheet 4: Data Flows
# ════════════════════════════════════════════════════════════════════════════
ws_df = new_sheet("Data Flows")
title_row(ws_df, "Data Flows — Step-by-Step Request Lifecycle for Key Operations", cols=7)
widths(ws_df, ("A", 6), ("B", 24), ("C", 20), ("D", 12), ("E", 35), ("F", 35), ("G", 30))

flows = [
    # ── Flow A: Create Formula ──────────────────────────────────────────────
    ("A", "CREATE FORMULA",  "", "", "", "", ""),
    ("A1", "User", "2 · UI", "All",
     "Fills ingredient amounts in FormulaTable; clicks Save",
     "Form input (ingredient IDs + amounts)",
     "Calls store action saveFormula()"),
    ("A2", "useFormulaStore", "3 · State", "All",
     "Dispatches POST /api/formulas/ via Axios",
     "{name, standard_id, items:[{ingredient, amount_kg}]}",
     "Awaits HTTP response"),
    ("A3", "Axios", "4 · API Client", "All",
     "Attaches Authorization header; sends HTTP POST",
     "Token header + JSON body",
     "Network request to backend"),
    ("A4", "StandaloneBypassAuth / TokenAuth", "5 · Auth", "All",
     "Authenticates request; attaches user to request.user",
     "Header or bypass",
     "request.user = local_admin (Desktop) or token user (Cloud)"),
    ("A5", "FormulaListView.perform_create()", "6 · API", "All",
     "Validates serializer; injects created_by=request.user",
     "Validated serializer data",
     "Calls serializer.save(created_by=user)"),
    ("A6", "FormulaSerializer.create()", "6 · API", "All",
     "Creates Formula + FormulaItem rows in ORM",
     "Validated data dict",
     "formula instance + items"),
    ("A7", "FormulaService.calculate_and_save()", "7 · BizLogic", "All",
     "Iterates items, computes total_cost, batch_size, cost_per_kg",
     "Formula instance with items",
     "Updated formula saved to DB"),
    ("A8", "Django ORM", "8 · ORM", "All",
     "INSERT Formula + FormulaItems; UPDATE Formula snapshot",
     "SQL INSERT / UPDATE",
     "Row IDs returned"),
    ("A9", "Database", "9 · DB", "All",
     "Persists rows to SQLite (Desktop) or PostgreSQL (Cloud)",
     "SQL",
     "Committed rows"),
    ("A10", "FormulaListView → response", "6 · API", "All",
     "Serializes saved formula; returns HTTP 201",
     "Formula + items queryset",
     "JSON with cost_per_kg, total_cost, items[]"),
    ("A11", "useFormulaStore", "3 · State", "All",
     "Updates local formula list; triggers reactivity",
     "API response JSON",
     "UI re-renders with new formula"),

    ("", "", "", "", "", "", ""),

    # ── Flow B: Cloud Login ─────────────────────────────────────────────────
    ("B", "CLOUD LOGIN (Token Auth)", "", "", "", "", ""),
    ("B1", "User", "2 · UI", "Cloud",
     "Enters email + password on login screen",
     "username, password",
     "Calls store action login()"),
    ("B2", "Axios", "4 · API Client", "Cloud",
     "POST /api/auth/token/login/ (Djoser endpoint)",
     "{username, password}",
     "HTTP POST to backend"),
    ("B3", "Djoser TokenCreateView", "5 · Auth", "Cloud",
     "Validates credentials; creates/returns DRF AuthToken",
     "username + password",
     "{auth_token: 'abc123...'}"),
    ("B4", "usePreferencesStore / useFormulaStore", "3 · State", "Cloud",
     "Stores token in memory; sets Axios Authorization header; fetches initial data",
     "auth_token",
     "Subsequent requests carry 'Token abc123...' header"),

    ("", "", "", "", "", "", ""),

    # ── Flow C: Desktop Startup ─────────────────────────────────────────────
    ("C", "DESKTOP STARTUP", "", "", "", "", ""),
    ("C1", "Tauri", "1 · Shell", "Desktop",
     "Launches beforeDevCommand: 'python run_api.py'",
     "OS process spawn",
     "feed_calc_api.exe (or python process) starts"),
    ("C2", "run_api.py", "9 · DB / 7 · BizLogic", "Desktop",
     "Runs migrate; calls _load_seed_fixture() if Ingredient.count()==0",
     "DB path from APPDATA",
     "DB ready; 347 ingredients seeded on first launch"),
    ("C3", "Waitress", "5 · Auth", "Desktop",
     "serve() starts listening on port 8042",
     "Django WSGI app",
     "startup.log: 'calling serve() on port 8042'"),
    ("C4", "App.vue loading screen", "2 · UI", "Desktop",
     "Polls GET /api/ingredients/?page_size=1 every 1 s until HTTP 200",
     "Unauthenticated GET",
     "StandaloneBypassAuth auto-responds 200 → loading screen dismissed"),
    ("C5", "useFormulaStore / useIngredientStore", "3 · State", "Desktop",
     "Fetches initial ingredient list and formula list",
     "GET /api/ingredients/ and /api/formulas/",
     "Stores populated; main UI rendered"),

    ("", "", "", "", "", "", ""),

    # ── Flow D: CSV Import ──────────────────────────────────────────────────
    ("D", "CSV IMPORT", "", "", "", "", ""),
    ("D1", "User", "2 · UI", "All",
     "Selects a previously exported CSV file via FormulaExporter upload",
     "File object (CSV bytes)",
     "Calls store action importFromCSV(file)"),
    ("D2", "Axios", "4 · API Client", "All",
     "POST /api/formulas/import/ as multipart/form-data",
     "FormData with 'file' field",
     "HTTP POST"),
    ("D3", "FormulaImportView.post()", "6 · API", "All",
     "Reads uploaded file bytes; calls _parse_formula_csv()",
     "Raw CSV bytes",
     "(meta dict, [(name, amount_kg)] list)"),
    ("D4", "_parse_formula_csv()", "7 · BizLogic", "All",
     "Strips BOM; parses meta rows and composition section; skips bad amounts",
     "CSV bytes",
     "meta={}, items=[(name, kg), ...]"),
    ("D5", "FormulaImportView — ingredient matching", "6 · API", "All",
     "For each name: exact match → iexact → icontains in public+own ingredients",
     "Ingredient names from CSV",
     "matched=[(id, kg)], unmatched=['name',...]"),
    ("D6", "FormulaService.calculate_and_save()", "7 · BizLogic", "All",
     "Creates Formula + FormulaItems; computes cost snapshot",
     "matched ingredient IDs + amounts",
     "Saved formula with cost fields"),
    ("D7", "FormulaImportView → response", "6 · API", "All",
     "Returns 201 with formula data and unmatched names list",
     "Serialized formula",
     "{formula: {...}, unmatched: ['name']}"),
]

FLOW_HDR_BG = {
    "A": "C00000", "B": "0070C0", "C": "375623", "D": "7030A0"
}

hdr(ws_df, 2, ["Step", "Actor", "Layer", "Platform", "Action", "Input", "Output"])
row = 3
for step_data in flows:
    step_id = step_data[0]
    ws_df.row_dimensions[row].height = 55
    is_section = step_data[1] in ("CREATE FORMULA", "CLOUD LOGIN (Token Auth)",
                                   "DESKTOP STARTUP", "CSV IMPORT")
    is_blank = all(v == "" for v in step_data)
    if is_section:
        letter = step_id
        bg_s = FLOW_HDR_BG.get(letter, "1F3864")
        ws_df.merge_cells(f"A{row}:G{row}")
        c = ws_df.cell(row=row, column=1,
                       value=f"  FLOW {step_id} — {step_data[1]}")
        c.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
        c.fill = fill(bg_s)
        c.alignment = left()
        c.border = thin()
        ws_df.row_dimensions[row].height = 24
    elif is_blank:
        ws_df.row_dimensions[row].height = 8
        for col in range(1, 8):
            c = ws_df.cell(row=row, column=col, value="")
            c.fill = fill("FFFFFF")
    else:
        row_bg = "F5F5F5" if row % 2 == 0 else "FFFFFF"
        for col, val in enumerate(step_data, 1):
            c = ws_df.cell(row=row, column=col, value=val)
            c.border = thin()
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if col == 1:
                c.font = Font(name=FONT, bold=True, size=10, color="1F1F1F")
                c.fill = fill("E8E8E8")
                c.alignment = center()
            elif col == 3:
                layer_letter = val[0] if val else ""
                layer_bg = {
                    "1": L["Shell"][0], "2": L["UI"][0], "3": L["State"][0],
                    "4": L["APIclient"][0], "5": L["Auth"][0], "6": L["Router"][0],
                    "7": L["BizLogic"][0], "8": L["ORM"][0], "9": L["DB"][0],
                }.get(layer_letter, "FFFFFF")
                layer_fg = {
                    "1": L["Shell"][1], "2": L["UI"][1], "3": L["State"][1],
                    "4": L["APIclient"][1], "5": L["Auth"][1], "6": L["Router"][1],
                    "7": L["BizLogic"][1], "8": L["ORM"][1], "9": L["DB"][1],
                }.get(layer_letter, "000000")
                c.font = Font(name=FONT, bold=True, size=9, color=layer_fg)
                c.fill = fill(layer_bg)
                c.alignment = center()
            else:
                c.font = Font(name=FONT, size=9, color="1F1F1F")
                c.fill = fill(row_bg)
    row += 1

ws_df.freeze_panes = "A3"

# ════════════════════════════════════════════════════════════════════════════
# Sheet 5: E2E Test Design
# ════════════════════════════════════════════════════════════════════════════
ws_e2e = new_sheet("E2E Test Design")
title_row(ws_e2e, "End-to-End Test Design — Scenarios, Tool, and Layer Coverage", cols=9)

hdr(ws_e2e, 2, ["Test ID", "Scenario", "Platform", "Tool",
               "Preconditions", "Steps", "Expected Result", "Layers Exercised", "Status"])
widths(ws_e2e,
    ("A", 9), ("B", 30), ("C", 12), ("D", 18),
    ("E", 30), ("F", 45), ("G", 40), ("H", 28), ("I", 14))

from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list",
    formula1='"Not Started,Pass,Fail,In Progress,Blocked,N/A"',
    allow_blank=True, showDropDown=False)
dv.sqref = "I3:I30"
ws_e2e.add_data_validation(dv)

STATUS_COLORS = {"Not Started":"D9D9D9","Pass":"C6EFCE","Fail":"FFC7CE",
                 "In Progress":"FFEB9C","Blocked":"FCE4D6","N/A":"FFFFFF"}

e2e_tests = [
    ("E2E-001", "Full formula creation golden path",
     "Cloud", "Playwright",
     "Clean DB; 2 ingredients seeded; 1 standard seeded; user account exists",
     "1. Navigate to app\n2. Log in with valid credentials\n3. Open New Formula\n"
     "4. Add 2 ingredients with amounts\n5. Set standard\n6. Click Save",
     "Formula appears in list with correct cost_per_kg; "
     "nutrient analysis panels populate without error",
     "L1→L2→L3→L4→L5→L6→L7→L8→L9", "Not Started"),

    ("E2E-002", "Nutrient analysis panel renders for saved formula",
     "Cloud", "Playwright",
     "Formula with corn + SBM saved; NRC swine grower standard linked",
     "1. Log in\n2. Open existing formula\n3. Navigate to Swine panel",
     "All nutrient rows visible; status badges (over/under/ok) correctly coloured; "
     "no console errors",
     "L1→L2→L3 (computed)", "Not Started"),

    ("E2E-003", "CSV export → import round-trip",
     "Cloud", "Playwright",
     "Formula with 3 ingredients exists; all 3 ingredients in DB",
     "1. Open formula\n2. Click Export CSV\n3. Navigate to Import\n"
     "4. Upload downloaded CSV\n5. Check imported formula",
     "Imported formula has same 3 ingredients and amounts; "
     "unmatched: [] in response; cost_per_kg within 0.01 of original",
     "L1→L2→L3→L4→L6→L7→L8→L9", "Not Started"),

    ("E2E-004", "Login → logout → access denied",
     "Cloud", "Playwright",
     "User account exists",
     "1. Log in\n2. Verify /api/auth/users/me/ returns 200\n3. Log out\n"
     "4. Attempt GET /api/formulas/",
     "After logout, /api/formulas/ returns 401; no stale token in store",
     "L1→L2→L3→L4→L5", "Not Started"),

    ("E2E-005", "Non-owner cannot edit another user's formula",
     "Cloud", "Playwright",
     "UserA's formula exists; logged in as UserB",
     "1. Log in as UserB\n2. Navigate to UserA's formula\n3. Attempt to click Save/Edit",
     "Edit controls hidden or return 403; formula data visible (read-only)",
     "L1→L2→L3→L4→L5→L6 (IsOwnerOrReadOnly)", "Not Started"),

    ("E2E-006", "Ingredient search and filter",
     "Cloud", "Playwright",
     "DB has 10+ public ingredients including one named '玉米'",
     "1. Log in\n2. Open ingredient browser\n3. Type '玉米' in search",
     "Only matching ingredients shown; debounce fires correctly; "
     "no duplicate API calls",
     "L1→L2→L3→L4→L6→L8", "Not Started"),

    ("E2E-007", "Desktop app startup and first-use seeding",
     "Desktop", "Manual / tauri-driver",
     "Fresh install — empty DB; feed_calc_api.exe available",
     "1. Launch desktop app\n2. Wait for loading screen to dismiss\n3. Open ingredient list",
     "Loading screen disappears within 30 s; ingredient list shows 347+ items "
     "(crawled_seed.json loaded); startup.log ends with 'calling serve() on port 8042'",
     "L1→L2→L3→L4→L5→L6→L8→L9 (seeding)", "Not Started"),

    ("E2E-008", "Desktop formula create without login",
     "Desktop", "Manual / tauri-driver",
     "Desktop app running; no login screen shown (StandaloneBypassAuth active)",
     "1. Launch desktop app\n2. Create new formula with 2 ingredients\n3. Save",
     "Formula saved as local_admin; no auth prompt shown; "
     "cost snapshot populated correctly",
     "L1→L2→L3→L4→L5 (bypass)→L6→L7→L8→L9", "Not Started"),

    ("E2E-009", "Preference mode toggle persists across sessions",
     "Cloud", "Playwright",
     "User logged in; UserPreference exists or will be created",
     "1. Toggle nutrient display to 'absolute'\n2. Reload page\n3. Check panel display mode",
     "After reload, display mode is still 'absolute'; "
     "GET /api/preferences/ returns nutrient_display_mode='absolute'",
     "L1→L2→L3→L4→L6→L8→L9", "Not Started"),

    ("E2E-010", "CSV import with partial ingredient match",
     "Cloud", "Playwright",
     "DB has 2 of 3 ingredient names from the CSV",
     "1. Upload CSV with 3 ingredients (1 unknown)\n2. Check import result",
     "HTTP 201 returned; formula created with 2 matched ingredients; "
     "UI displays warning listing the 1 unmatched name",
     "L1→L2→L3→L4→L6→L7→L8", "Not Started"),
]

for i, t in enumerate(e2e_tests, start=3):
    ws_e2e.row_dimensions[i].height = 80
    status = t[-1]
    s_bg = STATUS_COLORS.get(status, "FFFFFF")
    platform_colors = {"Cloud": ("DAEEF3", "17375E"), "Desktop": ("E2EFDA", "375623"),
                       "Both": ("FFF2CC", "7F6000")}
    p_bg, p_fg = platform_colors.get(t[2], ("FFFFFF", "000000"))
    for col, val in enumerate(t, 1):
        c = ws_e2e.cell(row=i, column=col, value=val)
        c.border = thin()
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if col == 1:
            c.font = Font(name=FONT, bold=True, size=10, color="1F1F1F")
            c.fill = fill("E8E8E8")
            c.alignment = center()
        elif col == 3:
            c.font = Font(name=FONT, bold=True, size=9, color=p_fg)
            c.fill = fill(p_bg)
            c.alignment = center()
        elif col == 9:
            c.font = Font(name=FONT, size=9, color="1F1F1F")
            c.fill = fill(s_bg)
            c.alignment = center()
        else:
            bg = "F5F5F5" if i % 2 == 0 else "FFFFFF"
            c.font = Font(name=FONT, size=9, color="1F1F1F")
            c.fill = fill(bg)

# ════════════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════════════
out = r"C:\202511\05_sideproject\feed_calc_project\feed_calc_architecture.xlsx"
wb.save(out)
print(f"Saved: {out}")

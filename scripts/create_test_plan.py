from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ─────────────────────────────────────────────
# Shared styles
# ─────────────────────────────────────────────
FONT_NAME = "Arial"

def hdr_font(bold=True, size=10, color="FFFFFF"):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color)

def body_font(bold=False, size=10, color="000000"):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color)

def solid_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_wrap():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

PRIORITY_COLORS = {
    "P0": ("C00000", "FFFFFF"),  # dark red, white text
    "P1": ("E26B0A", "FFFFFF"),  # dark orange
    "P2": ("7030A0", "FFFFFF"),  # purple
    "P3": ("0070C0", "FFFFFF"),  # blue
    "P4": ("375623", "FFFFFF"),  # dark green
    "P5": ("808080", "FFFFFF"),  # grey
}

STATUS_COLORS = {
    "Not Started": "D9D9D9",
    "Pass":        "C6EFCE",
    "Fail":        "FFC7CE",
    "In Progress": "FFEB9C",
    "Blocked":     "FCE4D6",
    "N/A":         "FFFFFF",
}

COL_HEADERS = [
    "Test ID", "Priority", "Area", "Sub-Area",
    "Test Case", "Preconditions", "Input / Action", "Expected Result",
    "Status", "Notes",
]
COL_WIDTHS = [10, 8, 18, 20, 40, 30, 40, 40, 14, 30]

def apply_header_row(ws, row=1, bg="1F3864", text_color="FFFFFF"):
    for col_idx, header in enumerate(COL_HEADERS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(name=FONT_NAME, bold=True, size=10, color=text_color)
        cell.fill = solid_fill(bg)
        cell.alignment = center()
        cell.border = thin_border()

def set_col_widths(ws):
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_status_validation(ws, first_data_row, last_data_row):
    dv = DataValidation(
        type="list",
        formula1='"Not Started,Pass,Fail,In Progress,Blocked,N/A"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = f"I{first_data_row}:I{last_data_row}"
    ws.add_data_validation(dv)

def write_test_row(ws, row, test_id, priority, area, sub_area,
                   test_case, preconditions, input_action, expected,
                   status="Not Started", notes=""):
    values = [test_id, priority, area, sub_area,
              test_case, preconditions, input_action, expected, status, notes]
    p_bg, p_fg = PRIORITY_COLORS.get(priority, ("FFFFFF", "000000"))
    s_bg = STATUS_COLORS.get(status, "FFFFFF")
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = body_font()
        cell.alignment = left_wrap()
        cell.border = thin_border()
        if col_idx == 2:  # Priority
            cell.fill = solid_fill(p_bg)
            cell.font = Font(name=FONT_NAME, bold=True, size=10, color=p_fg)
            cell.alignment = center()
        elif col_idx == 9:  # Status
            cell.fill = solid_fill(s_bg)
            cell.alignment = center()
        else:
            cell.fill = solid_fill("FFFFFF")
    ws.row_dimensions[row].height = 50

# ─────────────────────────────────────────────
# Sheet 1: Summary / Legend
# ─────────────────────────────────────────────
ws_sum = wb.active
ws_sum.title = "Summary"
ws_sum.sheet_view.showGridLines = False

def sum_cell(r, c, val, bold=False, size=11, bg=None, color="000000", h_align="left"):
    cell = ws_sum.cell(row=r, column=c, value=val)
    cell.font = Font(name=FONT_NAME, bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)
    cell.border = thin_border()
    if bg:
        cell.fill = solid_fill(bg)
    return cell

ws_sum.column_dimensions["A"].width = 22
ws_sum.column_dimensions["B"].width = 50
ws_sum.column_dimensions["C"].width = 15
ws_sum.column_dimensions["D"].width = 15

# Title
ws_sum.merge_cells("A1:D1")
t = ws_sum.cell(row=1, column=1, value="Feed Formula Calculator — Test Plan")
t.font = Font(name=FONT_NAME, bold=True, size=16, color="FFFFFF")
t.fill = solid_fill("1F3864")
t.alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 35

ws_sum.merge_cells("A2:D2")
ws_sum.cell(row=2, column=1, value="Version 1.0  |  Date: 2026-05-26  |  Author: Shokahsu")
ws_sum.cell(row=2, column=1).font = Font(name=FONT_NAME, italic=True, size=10, color="595959")
ws_sum.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[2].height = 20

# Priority legend
sum_cell(4, 1, "Priority Legend", bold=True, size=11, bg="1F3864", color="FFFFFF", h_align="center")
sum_cell(4, 2, "Description", bold=True, size=11, bg="1F3864", color="FFFFFF", h_align="center")
sum_cell(4, 3, "Sheet", bold=True, size=11, bg="1F3864", color="FFFFFF", h_align="center")
sum_cell(4, 4, "# Tests", bold=True, size=11, bg="1F3864", color="FFFFFF", h_align="center")

legend_data = [
    ("P0", "Core Business Logic — run on every commit", "P0 - Business Logic", 9),
    ("P1", "API Contract Tests — explicit scenario-based", "P1 - API Tests", 23),
    ("P2", "Permission & Auth Mode Tests", "P2 - Permissions", 6),
    ("P3", "Frontend Unit Tests (Vue / Vitest)", "P3 - Frontend", 8),
    ("P4", "Desktop Integration Tests (manual/semi-auto)", "P4 - Desktop", 6),
    ("P5", "Known Bug Regression Tests", "P5 - Regression", 3),
]

for i, (p, desc, sheet, count) in enumerate(legend_data, start=5):
    bg, fg = PRIORITY_COLORS[p]
    sum_cell(i, 1, p, bold=True, bg=bg, color=fg, h_align="center")
    sum_cell(i, 2, desc, bg="F2F2F2")
    sum_cell(i, 3, sheet, bg="F2F2F2")
    sum_cell(i, 4, count, bg="F2F2F2", h_align="center")
    ws_sum.row_dimensions[i].height = 22

# Status legend
sum_cell(12, 1, "Status Legend", bold=True, size=11, bg="1F3864", color="FFFFFF", h_align="center")
sum_cell(12, 2, "Meaning", bold=True, size=11, bg="1F3864", color="FFFFFF", h_align="center")
ws_sum.merge_cells("C12:D12")
ws_sum.row_dimensions[12].height = 22

for i, (status, meaning) in enumerate([
    ("Not Started", "Test case has not yet been executed"),
    ("In Progress", "Test is being written or executed"),
    ("Pass",        "Test executed and result matched expected"),
    ("Fail",        "Test executed but result did NOT match expected"),
    ("Blocked",     "Cannot execute — dependency or environment issue"),
    ("N/A",         "Not applicable in current context"),
], start=13):
    bg = STATUS_COLORS[status]
    sum_cell(i, 1, status, bg=bg, h_align="center")
    sum_cell(i, 2, meaning, bg="F2F2F2")
    ws_sum.merge_cells(f"C{i}:D{i}")
    ws_sum.row_dimensions[i].height = 20

# Scope note
ws_sum.merge_cells("A20:D20")
scope = ws_sum.cell(row=20, column=1,
    value="Scope: Django REST API (accounts, ingredients, standards, formulas), "
          "FormulaService calculation engine, CSV import/export, "
          "Desktop sidecar auth, Vue 3 composables. "
          "Analysis and reports apps are out-of-scope (placeholder stubs).")
scope.font = Font(name=FONT_NAME, size=9, italic=True, color="595959")
scope.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_sum.row_dimensions[20].height = 40

# ─────────────────────────────────────────────
# Helper: create a test-case sheet
# ─────────────────────────────────────────────
def make_sheet(title, bg_color):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    apply_header_row(ws, row=1, bg=bg_color)
    set_col_widths(ws)
    return ws

# ─────────────────────────────────────────────
# Sheet 2: P0 — Core Business Logic
# ─────────────────────────────────────────────
ws0 = make_sheet("P0 - Business Logic", PRIORITY_COLORS["P0"][0])

p0_tests = [
    # FormulaService
    ("P0-001", "P0", "FormulaService", "calculate_and_save",
     "Two ingredients with known prices → correct cost snapshot",
     "Two Ingredient objects with cost_per_kg_twd set; Formula with those items",
     "corn: 60 kg @ 8 TWD/kg, sbm: 40 kg @ 12 TWD/kg",
     "batch_size=100, total_cost=960, cost_per_kg=9.60"),

    ("P0-002", "P0", "FormulaService", "calculate_and_save",
     "Ingredient with cost_per_kg_twd=None treated as 0",
     "Ingredient with no cost set; Formula with that item",
     "ingredient: 50 kg, cost=None",
     "No exception raised; total_cost=0.0, cost_per_kg=0.0"),

    ("P0-003", "P0", "FormulaService", "calculate_and_save",
     "Empty formula (no items) → no division-by-zero",
     "Formula with zero FormulaItems",
     "formula.items is empty queryset",
     "batch_size=0, total_cost=0, cost_per_kg=0; no ZeroDivisionError"),

    ("P0-004", "P0", "FormulaService", "calculate_and_save",
     "cost_per_kg == total_cost / batch_size identity",
     "Formula with 3+ ingredients; all have valid costs",
     "Mixed weights and prices summing to total_weight > 0",
     "cost_per_kg == total_cost / batch_size to float precision"),

    # _parse_formula_csv
    ("P0-005", "P0", "CSV Parser", "_parse_formula_csv",
     "Valid BOM-prefixed CSV (as exported by FormulaExporter.vue) parsed correctly",
     "CSV bytes matching the export format with BOM prefix",
     "UTF-8 BOM + 配方名稱,TestFormula header + 原料名稱/使用量 section",
     "meta['name']='TestFormula'; items list contains correct (name, amount) tuples"),

    ("P0-006", "P0", "CSV Parser", "_parse_formula_csv",
     "Header row '原料名稱' is NOT included in items list",
     "Valid CSV with composition section header",
     "Row: 原料名稱,使用量(kg)",
     "items does not contain ('原料名稱', ...) entry"),

    ("P0-007", "P0", "CSV Parser", "_parse_formula_csv",
     "Non-numeric amount row silently skipped",
     "CSV composition section with one bad amount",
     "玉米,abc (unparseable float)",
     "No exception; that row absent from items; other rows present"),

    ("P0-008", "P0", "CSV Parser", "_parse_formula_csv",
     "Empty file → items is empty list",
     "Zero-byte or whitespace-only CSV",
     "b'' or b'  '",
     "meta={}, items=[]"),

    ("P0-009", "P0", "CSV Parser", "_parse_formula_csv",
     "Multi-byte CJK ingredient names preserved correctly",
     "CSV with Chinese ingredient names",
     "黃豆粕,40.0 in composition section",
     "items contains ('黃豆粕', 40.0); name not mangled"),
]

for r, t in enumerate(p0_tests, start=2):
    write_test_row(ws0, r, *t)

add_status_validation(ws0, 2, len(p0_tests) + 1)

# ─────────────────────────────────────────────
# Sheet 3: P1 — API Contract Tests
# ─────────────────────────────────────────────
ws1 = make_sheet("P1 - API Tests", PRIORITY_COLORS["P1"][0])

p1_tests = [
    # Auth
    ("P1-001", "P1", "Authentication", "Token Login",
     "Valid credentials → returns auth_token",
     "User exists in DB with known username/password",
     "POST /api/auth/token/login/ {username, password}",
     "HTTP 200; response body contains 'auth_token' key"),

    ("P1-002", "P1", "Authentication", "Token Login",
     "Wrong password → rejected",
     "User exists; provide incorrect password",
     "POST /api/auth/token/login/ {username, wrong_password}",
     "HTTP 400; no token in response"),

    ("P1-003", "P1", "Authentication", "Users Me",
     "Valid token → returns current user profile",
     "User logged in; valid token in Authorization header",
     "GET /api/auth/users/me/ with Token header",
     "HTTP 200; response contains 'username' field"),

    ("P1-004", "P1", "Authentication", "Users Me",
     "No token → 401",
     "No Authorization header",
     "GET /api/auth/users/me/ without token",
     "HTTP 401"),

    ("P1-005", "P1", "Authentication", "Token Logout",
     "Logout invalidates token",
     "User has valid token",
     "POST /api/auth/token/logout/ with Token header",
     "HTTP 204; subsequent GET /users/me/ returns 401"),

    # Ingredients
    ("P1-006", "P1", "Ingredients", "List",
     "Unauthenticated request → 401",
     "No Authorization header",
     "GET /api/ingredients/",
     "HTTP 401"),

    ("P1-007", "P1", "Ingredients", "List",
     "Authenticated → returns public + own ingredients paginated",
     "User authenticated; public ingredient + private ingredient of user exist",
     "GET /api/ingredients/ with valid token",
     "HTTP 200; pagination envelope; both ingredients present"),

    ("P1-008", "P1", "Ingredients", "Visibility",
     "Private ingredient NOT visible to other users",
     "UserA creates ingredient with is_public=False",
     "GET /api/ingredients/ as UserB",
     "UserA's private ingredient absent from results"),

    ("P1-009", "P1", "Ingredients", "Permissions",
     "User cannot DELETE another user's ingredient",
     "Ingredient owned by UserA",
     "DELETE /api/ingredients/{id}/ as UserB",
     "HTTP 403 or 404"),

    ("P1-010", "P1", "Ingredients", "Search",
     "Search filter returns matching ingredients",
     "Ingredient named '玉米' exists",
     "GET /api/ingredients/?search=玉米",
     "HTTP 200; results contain '玉米'; non-matching ingredients absent"),

    # Formulas CRUD
    ("P1-011", "P1", "Formulas", "Create",
     "Valid payload → 201 and cost computed",
     "User authenticated; ingredients and standard exist",
     "POST /api/formulas/ {name, standard_id, items:[{ingredient, amount_kg}]}",
     "HTTP 201; response includes cost_per_kg, total_cost, batch_size with correct values"),

    ("P1-012", "P1", "Formulas", "Read",
     "Owner retrieves formula with items",
     "Formula created by this user",
     "GET /api/formulas/{id}/",
     "HTTP 200; 'items' array present; ingredient IDs match"),

    ("P1-013", "P1", "Formulas", "Read",
     "Non-owner can read formula (read-only permission)",
     "Formula owned by UserA",
     "GET /api/formulas/{id}/ as UserB",
     "HTTP 200 (shared read access)"),

    ("P1-014", "P1", "Formulas", "Update",
     "Non-owner cannot update formula",
     "Formula owned by UserA",
     "PUT /api/formulas/{id}/ as UserB",
     "HTTP 403"),

    ("P1-015", "P1", "Formulas", "Delete",
     "Owner can delete their formula",
     "Formula owned by this user",
     "DELETE /api/formulas/{id}/",
     "HTTP 204; subsequent GET returns 404"),

    ("P1-016", "P1", "Formulas", "Delete",
     "Non-owner cannot delete formula",
     "Formula owned by UserA",
     "DELETE /api/formulas/{id}/ as UserB",
     "HTTP 403"),

    ("P1-017", "P1", "Formulas", "Cascade",
     "Formula with deleted standard → standard becomes null, formula survives",
     "Formula linked to a NutrientStandard; standard is deleted",
     "DELETE standard, then GET /api/formulas/{id}/",
     "HTTP 200; 'standard': null in response"),

    # Formula Import
    ("P1-018", "P1", "Formula Import", "CSV Upload",
     "Valid CSV → 201 with matched ingredients",
     "Ingredients exist in DB matching CSV names; user authenticated",
     "POST /api/formulas/import/ multipart with valid CSV file",
     "HTTP 201; formula created; 'unmatched': []"),

    ("P1-019", "P1", "Formula Import", "CSV Upload",
     "CSV with one unknown ingredient → partial match, 201",
     "Most ingredients matched; one name has no DB match",
     "POST /api/formulas/import/ CSV with one unknown ingredient",
     "HTTP 201; formula created; 'unmatched': ['UnknownName']"),

    ("P1-020", "P1", "Formula Import", "CSV Upload",
     "All ingredients unmatched → 400",
     "No ingredients in DB match any CSV name",
     "POST /api/formulas/import/ CSV with entirely unknown ingredient names",
     "HTTP 400; error message present; no formula created"),

    ("P1-021", "P1", "Formula Import", "CSV Upload",
     "No file attached → 400",
     "User authenticated",
     "POST /api/formulas/import/ with no file field",
     "HTTP 400; '未提供 CSV 檔案' error"),

    # Standards
    ("P1-022", "P1", "Standards", "Requirements",
     "Requirements list endpoint returns data",
     "NutrientStandard records exist",
     "GET /api/standards/requirements/ with valid token",
     "HTTP 200; list of standard objects"),

    ("P1-023", "P1", "Standards", "Stages",
     "Stages list endpoint returns data",
     "User authenticated",
     "GET /api/standards/stages/ with valid token",
     "HTTP 200; list of stage choices"),
]

for r, t in enumerate(p1_tests, start=2):
    write_test_row(ws1, r, *t)

add_status_validation(ws1, 2, len(p1_tests) + 1)

# ─────────────────────────────────────────────
# Sheet 4: P2 — Permissions & Auth Mode
# ─────────────────────────────────────────────
ws2 = make_sheet("P2 - Permissions", PRIORITY_COLORS["P2"][0])

p2_tests = [
    ("P2-001", "P2", "Desktop Auth", "StandaloneBypassAuth",
     "Request with no auth header auto-authenticated as local_admin",
     "Django running with settings/desktop.py; no Authorization header",
     "GET /api/ingredients/ with no Authorization header (desktop settings)",
     "HTTP 200; user is 'local_admin'; user.is_superuser=True"),

    ("P2-002", "P2", "Desktop Auth", "StandaloneBypassAuth",
     "local_admin can create formula without token",
     "Desktop settings; no auth header",
     "POST /api/formulas/ valid payload, no Authorization header",
     "HTTP 201; formula created_by = local_admin"),

    ("P2-003", "P2", "Desktop Auth", "StandaloneBypassAuth",
     "local_admin can delete any user's formula",
     "Desktop settings; formula owned by another user",
     "DELETE /api/formulas/{id}/ no auth header",
     "HTTP 204; formula deleted"),

    ("P2-004", "P2", "Cloud Auth", "Token isolation",
     "'Bearer standalone-admin' header on cloud settings does NOT bypass auth",
     "Django running with settings/base.py or prod.py",
     "GET /api/ingredients/ with 'Authorization: Bearer standalone-admin'",
     "HTTP 401 or 403 — bypass NOT granted in cloud mode"),

    ("P2-005", "P2", "User Preferences", "Upsert",
     "POST preferences creates record if none exists",
     "User authenticated; no UserPreference exists for user",
     "POST /api/preferences/ {nutrient_display_mode: 'percent'}",
     "HTTP 200 or 201; UserPreference created with nutrient_display_mode='percent'"),

    ("P2-006", "P2", "User Preferences", "Upsert",
     "Second POST updates, not duplicates, preference record",
     "User authenticated; UserPreference already exists",
     "POST /api/preferences/ {nutrient_display_mode: 'absolute'}",
     "HTTP 200; existing record updated; no duplicate row created"),
]

for r, t in enumerate(p2_tests, start=2):
    write_test_row(ws2, r, *t)

add_status_validation(ws2, 2, len(p2_tests) + 1)

# ─────────────────────────────────────────────
# Sheet 5: P3 — Frontend Unit Tests
# ─────────────────────────────────────────────
ws3 = make_sheet("P3 - Frontend", PRIORITY_COLORS["P3"][0])

p3_tests = [
    ("P3-001", "P3", "useGroupStatus", "Composable",
     "Over-fed nutrient → status: 'over'",
     "Vitest environment; composable imported",
     "Input: [{current: 10, target: 8, isMax: false}]",
     "Returns {status: 'over', count: 1}"),

    ("P3-002", "P3", "useGroupStatus", "Composable",
     "Deficient nutrient → status: 'under'",
     "Vitest environment; composable imported",
     "Input: [{current: 5, target: 8, isMax: false}]",
     "Returns {status: 'under', count: 1}"),

    ("P3-003", "P3", "useGroupStatus", "Composable",
     "Max-type constraint: value above target → status: 'over'",
     "Vitest environment",
     "Input: [{current: 15, target: 10, isMax: true}]",
     "Returns {status: 'over', count: 1}"),

    ("P3-004", "P3", "useGroupStatus", "Composable",
     "Empty input array → ok with count 0",
     "Vitest environment",
     "Input: []",
     "Returns {status: 'ok', count: 0}"),

    ("P3-005", "P3", "useFormulaStore", "Store",
     "Loading a formula populates items array correctly",
     "Pinia test setup; mock API returns formula with 2 items",
     "store.loadFormula(id) called with mocked response",
     "store.items has 2 entries with correct ingredient IDs and amounts"),

    ("P3-006", "P3", "useFormulaStore", "Store",
     "Object.assign pattern does not throw on saveForm update",
     "Pinia test setup; saveForm is reactive({})",
     "Object.assign(saveForm, newData) called",
     "No TypeError; saveForm fields updated correctly (regression: const-assignment bug)"),

    ("P3-007", "P3", "FormulaExporter", "CSV Export",
     "Downloaded CSV contains BOM prefix for Excel compatibility",
     "Component mounted with a formula; downloadCSV() triggered",
     "Trigger downloadCSV() action",
     "Blob content starts with UTF-8 BOM (\\xEF\\xBB\\xBF); 配方名稱 row present"),

    ("P3-008", "P3", "FormulaExporter", "CSV Export",
     "CSV round-trip: export then import returns same ingredient list",
     "Formula with known ingredients; _parse_formula_csv() available",
     "Export formula → parse the CSV bytes → compare items list",
     "All ingredient names and amounts match original formula items"),
]

for r, t in enumerate(p3_tests, start=2):
    write_test_row(ws3, r, *t)

add_status_validation(ws3, 2, len(p3_tests) + 1)

# ─────────────────────────────────────────────
# Sheet 6: P4 — Desktop Integration
# ─────────────────────────────────────────────
ws4 = make_sheet("P4 - Desktop", PRIORITY_COLORS["P4"][0])

p4_tests = [
    ("P4-001", "P4", "Sidecar Startup", "Port",
     "Sidecar listens on port 8042, not 8000",
     "Desktop build (.exe) launched fresh",
     "Check startup.log for port; try curl http://127.0.0.1:8042/api/ingredients/",
     "startup.log ends with 'calling serve() on port 8042'; HTTP 200 from 8042; 8000 not in use"),

    ("P4-002", "P4", "Sidecar Startup", "Seed on empty DB",
     "First launch with empty DB → fixture seeded automatically",
     "Delete or replace db.sqlite3 with empty file; launch app",
     "Launch feed_calc_api.exe; wait for ready; query ingredient count",
     "Ingredient.objects.count() > 0 (crawled_seed.json loaded)"),

    ("P4-003", "P4", "Sidecar Startup", "Seed skip on existing DB",
     "Second launch skips seeding (no duplicate error)",
     "DB already populated from previous run",
     "Restart sidecar; check startup.log and ingredient count",
     "No 'UNIQUE constraint' error in log; ingredient count unchanged"),

    ("P4-004", "P4", "Build Process", "Stale process",
     "Build fails with OS error 5 if feed_calc_api.exe still running",
     "feed_calc_api.exe running from previous dev session",
     "Attempt 'tauri build' without stopping sidecar",
     "OS error 5 (access denied); resolved by Stop-Process -Name feed_calc_api -Force"),

    ("P4-005", "P4", "Build Process", "PyInstaller",
     "djoser.urls guard prevents AppRegistryNotReady crash",
     "Desktop settings; sys.frozen=True equivalent",
     "Tauri build completes; launch built .exe",
     "App starts without AppRegistryNotReady traceback in startup.log"),

    ("P4-006", "P4", "CORS", "x-client-mode header",
     "Custom x-client-mode header not blocked by CORS",
     "Frontend sending x-client-mode header; desktop settings",
     "Preflight OPTIONS request with x-client-mode in Access-Control-Request-Headers",
     "CORS preflight returns 200; x-client-mode listed in Access-Control-Allow-Headers"),
]

for r, t in enumerate(p4_tests, start=2):
    write_test_row(ws4, r, *t)

add_status_validation(ws4, 2, len(p4_tests) + 1)

# ─────────────────────────────────────────────
# Sheet 7: P5 — Known Bug Regression Tests
# ─────────────────────────────────────────────
ws5 = make_sheet("P5 - Regression", PRIORITY_COLORS["P5"][0])

p5_tests = [
    ("P5-001", "P5", "Ingredient Model", "phytate_P_g_per_kg",
     "[EXPECTED FAIL] phytate_P field persisted to DB",
     "Bug: ingredients/models.py uses ':' instead of '=' for phytate_P_g_per_kg field assignment",
     "Create Ingredient(phytate_P_g_per_kg=1.5); reload from DB",
     "CURRENTLY FAILS: field is missing from DB schema. Fix: change ':' to '=' on model line 87, run migration"),

    ("P5-002", "P5", "Auth Classes", "Dual auth consistency",
     "DesktopStandaloneAuthentication and StandaloneBypassAuth behave consistently",
     "Known inconsistency: accounts/authentication.py vs config/settings/desktop.py define different bypass classes",
     "Run both auth paths; compare user created (desktop_default_user vs local_admin) and permission level",
     "Both should produce the same user identity and superuser status — currently they differ"),

    ("P5-003", "P5", "Ingredient Import", "cost_per_kg NOT NULL",
     "import_ingredients management command skips rows without crashing the overall import",
     "feed_database.xlsx has rows missing cost_per_kg_twd; NOT NULL constraint pre-existing",
     "Run 'python manage.py import_ingredients'; check error count vs success count",
     "Command completes; partial import succeeds; NOT NULL rows logged as warnings, not fatal errors"),
]

for r, t in enumerate(p5_tests, start=2):
    write_test_row(ws5, r, *t)

add_status_validation(ws5, 2, len(p5_tests) + 1)

# Mark P5-001 and P5-002 as "Blocked" (known bugs)
for row in [2, 3]:
    cell = ws5.cell(row=row, column=9)
    cell.value = "Blocked"
    cell.fill = solid_fill(STATUS_COLORS["Blocked"])

# ─────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────
out_path = r"C:\202511\05_sideproject\feed_calc_project\feed_calc_test_plan.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
